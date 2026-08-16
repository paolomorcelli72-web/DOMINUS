import streamlit as st
import pandas as pd
from openpyxl import load_workbook

FILE_XLSX = "DOMINUS 2026 DEFINITIVO.xlsx"

st.set_page_config(
    page_title="DOMINUS WEB",
    layout="wide"
)

st.title("📊 DOMINUS WEB")


@st.cache_resource
def load_wb():
    return load_workbook(
        FILE_XLSX,
        data_only=True
    )


wb = load_wb()

if "sheets" not in st.session_state:

    st.session_state.sheets = {}

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        data = []

        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        df = pd.DataFrame(data)
        df = df.fillna("")

        st.session_state.sheets[sheet] = df


sheet_name = st.sidebar.selectbox(
    "Foglio",
    wb.sheetnames
)

EDITABILI = [
    "ANAGRAFICA",
    "ASSETTO",
    "PATRIMONIO",
    "VALORE",
    "CUSTODIA",
    "LEADER"
]

df = st.session_state.sheets[sheet_name].copy()

for col in df.columns:
    df[col] = df[col].astype(str)

if sheet_name in EDITABILI:

    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="fixed",
        key=f"editor_{sheet_name}"
    )

else:

    edited_df = df

    st.dataframe(
        df,
        use_container_width=True
    )

st.session_state.sheets[sheet_name] = edited_df


def calcola_ambito(df):

    cg_tot = 0
    cg_no = 0

    for _, row in df.iterrows():

        valori = [str(v).strip().upper() for v in row.tolist()]

        risposta = None
        cg = 0

        for i, v in enumerate(valori):

            if v in ["SI", "NO"]:

                risposta = v

                try:
                    cg = float(
                        str(row.iloc[i + 1]).replace(",", ".")
                    )
                except:
                    cg = 0

                break

        if risposta is None:
            continue

        cg_tot += cg

        if risposta == "NO":
            cg_no += cg

    if cg_tot == 0:
        return 0

    return cg_no / cg_tot


try:

    assetto = calcola_ambito(
        st.session_state.sheets["ASSETTO"]
    )

    patrimonio = calcola_ambito(
        st.session_state.sheets["PATRIMONIO"]
    )

    valore = calcola_ambito(
        st.session_state.sheets["VALORE"]
    )

    custodia = calcola_ambito(
        st.session_state.sheets["CUSTODIA"]
    )

    dominus_score = round(
        assetto * 50 +
        patrimonio * 10 +
        valore * 30 +
        custodia * 10,
        2
    )

    st.divider()

    st.subheader("📈 DOMINUS SCORE LIVE")

    c1, c2, c3, c4, c5 = st.columns(5)

    c1.metric("Assetto", f"{assetto * 100:.2f}%")
    c2.metric("Patrimonio", f"{patrimonio * 100:.2f}%")
    c3.metric("Valore", f"{valore * 100:.2f}%")
    c4.metric("Custodia", f"{custodia * 100:.2f}%")
    c5.metric("Score", f"{dominus_score:.2f}")

    if dominus_score < 35:
        rating = "AAA"
    elif dominus_score < 43:
        rating = "AA"
    elif dominus_score < 50:
        rating = "A"
    elif dominus_score < 58:
        rating = "BBB"
    elif dominus_score < 65:
        rating = "BB"
    elif dominus_score < 80:
        rating = "B"
    else:
        rating = "D"

    st.success(f"🏆 Rating: {rating}")

except Exception as e:

    st.error(
        f"Errore calcolo score: {e}"
    )


if st.button("💾 Salva Workbook"):

    try:

        wb_save = load_workbook(
            FILE_XLSX,
            data_only=False
        )

        for foglio, dataframe in st.session_state.sheets.items():

            if foglio not in wb_save.sheetnames:
                continue

            ws = wb_save[foglio]

            max_r = min(
                len(dataframe),
                ws.max_row
            )

            max_c = min(
                len(dataframe.columns),
                ws.max_column
            )

            for r in range(max_r):

                for c in range(max_c):

                    try:

                        cella = ws.cell(
                            row=r + 1,
                            column=c + 1
                        )

                        valore = dataframe.iat[r, c]

                        if pd.isna(valore):
                            valore = None

                        try:
                            cella.value = valore
                        except:
                            pass

                    except:
                        pass

        wb_save.save(FILE_XLSX)

        st.success("✅ Workbook salvato")

    except Exception as e:

        st.error(
            f"Errore salvataggio: {e}"
        )
